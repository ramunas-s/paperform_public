import os
from pathlib import Path
import pandas as pd

from openpyxl.utils import get_column_letter

from labrep_recognizer.labrep_recognize_request import IOFileType, LabrepRecognizeRequest
from labrep_recognizer.normalization.normalization_survey_burnout import normalize_survey_burnout
from labrep_recognizer.shared.pdf_parser_logging import get_logger
from labrep_recognizer.shared.recognizer_cache import RecognizerCache
from labrep_recognizer.recognizer_infrastructure import RecognizerInfrastructure
from labrep_recognizer.shared.utils import make_dirs, file_to_sha256

log = get_logger(__name__)


def process_single_pdf_in_gs(uploaded_file_uri, recognizer_infrastructure):

    SAFE_EXECUTION = False

    input_files = single_pdf_gs_to_input_files_param(uploaded_file_uri)

    labrep_recognize_request = LabrepRecognizeRequest(
        recognizer_infrastructure=recognizer_infrastructure,
        input_files=input_files,
        parallel_execution=True,
    )

    if SAFE_EXECUTION:
        try:
            recognition_status_ok, recognition_error = labrep_recognize_request.recognize()

        except IndexError:
            recognition_status_ok = False
            recognition_error = f"IndexError "

        except TypeError:
            recognition_status_ok = False
            recognition_error = f"TypeError "
    else:
        recognition_status_ok, recognition_error = labrep_recognize_request.recognize()

    if recognition_status_ok:
        if recognizer_infrastructure._cache:
            recognizer_infrastructure._cache.persist_cache()
        return (
            True,
            recognition_error,
            labrep_recognize_request.df_header,
            labrep_recognize_request.df_details,
        )
    else:

        # TODO clarify why this fails?
        # recognizer_infrastructure._cache.persist_cache()

        return (
            False,
            recognition_error,
            None,
            None,
        )


def initialize_infrastructure():
    cache_file = os.environ.get("recognizer_casche")
    if cache_file is None or cache_file == "":
        cache = None
    else:
        cache = RecognizerCache(cache_file, 0)

    recognizer_infrastructure = RecognizerInfrastructure(
        project_id=os.environ.get("recognizer_project_id"),
        google_application_credentials=os.environ.get("recognizer_google_application_credentials"),
        ocr_abbyy_fr_engine_url=os.environ.get("ocr_abbyy_fr_engine_url"),
        recognizer_cache=cache,
    )
    recognizer_infrastructure.warm_up()
    return recognizer_infrastructure


def single_pdf_gs_to_input_files_param(uploaded_file_uri):
    input_files = [
        {
            "FILE_TYPE": IOFileType.INPUT_GS_PDF_RAW_LAB_REPORT,
            "FILE_PATH": uploaded_file_uri,
        }
    ]
    return input_files


def upload_pdf_to_gs(locac_pdf_path, recognizer_infrastructure: RecognizerInfrastructure):
    sha_256 = file_to_sha256(locac_pdf_path)
    target_file_name = sha_256 + ".pdf"
    uploaded_file_uri = recognizer_infrastructure.upload_pdf_to_google_bucket(locac_pdf_path, target_file_name)
    return uploaded_file_uri


def recognize_single_local_pdf(raw_pdf):
    recognizer_infrastructure = initialize_infrastructure()
    uploaded_file_uri = upload_pdf_to_gs(raw_pdf, recognizer_infrastructure)
    log.info(f"uploaded_file_uri: {uploaded_file_uri}")
    recognition_status_ok, recognition_error, df_header, df_details = process_single_pdf_in_gs(
        uploaded_file_uri, recognizer_infrastructure
    )
    return recognition_status_ok, recognition_error, df_header, df_details


def process_batch(input_dir, output_dir):
    input_pdfs = list(Path(input_dir).rglob("*.pdf"))

    status = []

    for input_pdf in input_pdfs:
        print(f"Testing: {input_pdf}")
        recognition_status_ok, recognition_error, df_header, df_details = recognize_single_local_pdf(input_pdf)
        status.append((input_pdf, recognition_status_ok, recognition_error))
        if recognition_status_ok:

            df_details_test_format = df_details[
                [
                    "test_type",
                    "test_name",
                    "units",
                    "value",
                ]
            ].copy()

            df_header = df_header.rename(
                columns={
                    "Laboratorija": "laboratory_name",
                    "Užsakovas": "requestor_organization_name",
                    "Gydytojas": "requestor_doctor_name",
                    "Pacientas": "patient_name",
                    "Gimimo data": "patient_birth_date",
                    "Lytis": "patient_gender",
                    "Mėginys paimtas": "test_date",
                    "Nr": "laboratory_internal_test_id",
                    "labrep_type": "labrep_type",
                    "laboratory_id": "laboratory_id",
                }
            )

            df_header_test_format = df_header[
                [
                    "laboratory_name",
                    "test_date",
                    "patient_name",
                    "patient_birth_date",
                    "patient_gender",
                ]
            ].copy()
            df_header_test_format["comment"] = ""

            df_header_test_format["test_date"] = pd.to_datetime(df_header_test_format["test_date"], errors="coerce")
            df_header_test_format["patient_birth_date"] = pd.to_datetime(
                df_header_test_format["patient_birth_date"], errors="coerce"
            )

            df_survey_burnout = normalize_survey_burnout(df_header_test_format, df_details_test_format)

            pd.options.display.max_rows = 300
            pd.options.display.width = 1000
            pd.options.display.max_columns = None
            pd.options.display.max_colwidth = 200

            print(df_header_test_format.T)
            print(df_survey_burnout)
            print(df_details_test_format)

            file_name = ".".join(str(input_pdf).split("/")[-1].split(".")[:-1])
            print(file_name)
            output_file_with_path = os.path.join(output_dir, f"{file_name}.xlsx")
            make_dirs(output_file_with_path)

            # TODO refactor with column withs, reuse in step4
            excel_writer = pd.ExcelWriter(output_file_with_path)
            df_header_test_format.to_excel(excel_writer, sheet_name="header", index=None)
            df_survey_burnout.to_excel(excel_writer, sheet_name="survey_burnout", index=None)
            df_details_test_format.to_excel(excel_writer, sheet_name="details", index=None)

            worksheet = excel_writer.sheets["header"]
            worksheet.column_dimensions[get_column_letter(1)].width = 30
            worksheet.column_dimensions[get_column_letter(2)].width = 20
            worksheet.column_dimensions[get_column_letter(3)].width = 30
            worksheet.column_dimensions[get_column_letter(4)].width = 15
            worksheet.column_dimensions[get_column_letter(5)].width = 15
            worksheet.column_dimensions[get_column_letter(6)].width = 60

            excel_writer.save()

        else:
            print(recognition_error)

    print("\n".join([str(element) for element in status]))
    print(f"Total: {len(status)}")
    print(f"Errors: {sum([not(element[1]) for element in status])}")


def main():

    # input_dir = "/Users/ramunas/projects/recognizer/temp/problem_2022-07-12"
    # output_dir = "/Users/ramunas/projects/recognizer/temp/problem_2022-07-12_result"


    # input_dir = "/Users/ramunas/projects/recognizer/paperform/data/survey/auto_recognition_pdf"
    # output_dir = "/Users/ramunas/projects/recognizer/paperform/data/survey/auto_recognized"


    input_dir = "/Users/ramunas/projects/recognizer/temp/problem_synlab_2022-07-13"
    output_dir = "/Users/ramunas/projects/recognizer/temp/problem_synlab_2022-07-13_result"


    process_batch(input_dir, output_dir)


if __name__ == "__main__":
    main()
